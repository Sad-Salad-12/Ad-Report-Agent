"""Read a weekly ZIP and classify source tables by their header fingerprints."""

from __future__ import annotations

import csv
import hashlib
import io
import re
import unicodedata
import warnings
import zipfile
from datetime import date, datetime
from pathlib import Path, PurePosixPath
from typing import Any, Dict, Iterable, List, Sequence, Tuple

from openpyxl import load_workbook

from .models import IngestedBundle, SourceKind, SourceRow, SourceTable


class IngestionError(ValueError):
    """Raised when the input bundle cannot be classified without guessing."""


def normalize_header(value: Any) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).replace("\ufeff", "")
    return re.sub(r"\s+", " ", text).strip().casefold()


HEADER_FINGERPRINTS: Dict[str, frozenset[str]] = {
    "overall": frozenset(
        normalize_header(item)
        for item in (
            "Starts",
            "Ends",
            "Amount spent (USD)",
            "Purchases conversion value",
            "Cost per purchase",
            "Link clicks",
            "Reporting starts",
            "Reporting ends",
        )
    ),
    "by_day": frozenset(
        normalize_header(item)
        for item in (
            "Day",
            "Adds to cart",
            "Purchases",
            "Purchase ROAS (return on ad spend)",
            "Amount spent (USD)",
            "Purchases conversion value",
        )
    ),
    "by_product": frozenset(
        normalize_header(item)
        for item in (
            "Campaign name",
            "Name",
            "Average purchases conversion value",
            "Purchases conversion value",
        )
    ),
    "campaign": frozenset(
        normalize_header(item)
        for item in (
            "Campaign name",
            "Purchases rate per landing page views",
            "Add to cart Rate",
            "Cost per purchase",
        )
    ),
    "creative": frozenset(
        normalize_header(item)
        for item in (
            "Ads",
            "Campaign name",
            "Cost per add to cart",
            "Purchase ROAS (return on ad spend)",
        )
    ),
    "traffic_campaign": frozenset(
        normalize_header(item)
        for item in (
            "Campaign name",
            "Landing page views",
            "Cost per landing page view",
            "Landing page views rate per link clicks",
        )
    ),
    "audience": frozenset(
        normalize_header(item)
        for item in (
            "Ad Set Name",
            "Adds to cart",
            "Purchases conversion value",
            "Amount spent (USD)",
        )
    ),
    "keyword": frozenset(
        normalize_header(item)
        for item in ("Keyword", "Conversions", "Conv. value / cost", "Cost", "Conv. value")
    ),
}

REQUIRED_SOURCE_KINDS = frozenset(HEADER_FINGERPRINTS)
SUPPORTED_SOURCE_SUFFIXES = frozenset({".xlsx", ".csv", ".tsv"})


def classify_headers(headers: Sequence[Any]) -> SourceKind:
    """Return the one report kind whose required header set is present."""

    normalized = {normalize_header(header) for header in headers if normalize_header(header)}
    matches = [kind for kind, required in HEADER_FINGERPRINTS.items() if required <= normalized]
    if len(matches) != 1:
        details = ", ".join(sorted(normalized))
        if not matches:
            raise IngestionError(f"unknown header fingerprint: [{details}]")
        raise IngestionError(f"ambiguous header fingerprint {matches}: [{details}]")
    return matches[0]  # type: ignore[return-value]


def _json_safe(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def _is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _date_value(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        match = re.search(r"20\d{2}-\d{2}-\d{2}", value)
        if match:
            return date.fromisoformat(match.group(0))
    return None


def _header_index(headers: Sequence[str], wanted: str) -> int | None:
    wanted_normalized = normalize_header(wanted)
    for index, header in enumerate(headers):
        if normalize_header(header) == wanted_normalized:
            return index
    return None


def _period(headers: Sequence[str], rows: Sequence[SourceRow]) -> Tuple[date | None, date | None]:
    start_index = _header_index(headers, "Reporting starts")
    end_index = _header_index(headers, "Reporting ends")
    if start_index is None or end_index is None:
        return None, None
    starts: List[date] = []
    ends: List[date] = []
    start_header = headers[start_index]
    end_header = headers[end_index]
    for row in rows:
        start = _date_value(row.values.get(start_header))
        end = _date_value(row.values.get(end_header))
        if start:
            starts.append(start)
        if end:
            ends.append(end)
    return (min(starts) if starts else None, max(ends) if ends else None)


def _build_table(
    *,
    file_name: str,
    sheet_name: str,
    header_row_number: int,
    raw_headers: Sequence[Any],
    data_rows: Iterable[Tuple[int, Sequence[Any]]],
) -> SourceTable:
    last_header = max((index for index, value in enumerate(raw_headers) if not _is_blank(value)), default=-1)
    if last_header < 0:
        raise IngestionError(f"{file_name}: header row is blank")
    headers = [str(value).strip() if not _is_blank(value) else f"__blank_{index + 1}" for index, value in enumerate(raw_headers[: last_header + 1])]
    normalized = [normalize_header(header) for header in headers if not header.startswith("__blank_")]
    if len(normalized) != len(set(normalized)):
        raise IngestionError(f"{file_name}: duplicate headers are not supported")
    kind = classify_headers(headers)
    rows: List[SourceRow] = []
    for row_number, raw_row in data_rows:
        values = list(raw_row[: len(headers)])
        values.extend([None] * (len(headers) - len(values)))
        if all(_is_blank(value) for value in values):
            continue
        rows.append(
            SourceRow(
                row_number=row_number,
                values={header: _json_safe(value) for header, value in zip(headers, values)},
            )
        )
    if not rows:
        raise IngestionError(f"{file_name}: classified as {kind} but contains no data rows")
    period_start, period_end = _period(headers, rows)
    return SourceTable(
        kind=kind,
        file_name=file_name,
        sheet_name=sheet_name,
        header_row_number=header_row_number,
        headers=headers,
        rows=rows,
        period_start=period_start,
        period_end=period_end,
    )


def _read_xlsx(file_name: str, payload: bytes) -> SourceTable:
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message="Workbook contains no default style")
        workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    try:
        candidates: List[SourceTable] = []
        failures: List[str] = []
        for worksheet in workbook.worksheets:
            # Do not trust max_row/UsedRange: exported files can carry formatting
            # to the bottom of a sheet. A long blank tail is not report data.
            materialized: List[Sequence[Any]] = []
            blank_streak = 0
            for raw_row in worksheet.iter_rows(values_only=True):
                materialized.append(raw_row)
                if all(_is_blank(value) for value in raw_row):
                    blank_streak += 1
                else:
                    blank_streak = 0
                if len(materialized) > 20 and blank_streak >= 100:
                    break
            for offset, raw_row in enumerate(materialized[:20], start=1):
                if sum(not _is_blank(value) for value in raw_row) < 2:
                    continue
                try:
                    classify_headers(raw_row)
                except IngestionError as exc:
                    failures.append(str(exc))
                    continue
                table = _build_table(
                    file_name=file_name,
                    sheet_name=worksheet.title,
                    header_row_number=offset,
                    raw_headers=raw_row,
                    data_rows=((index, row) for index, row in enumerate(materialized[offset:], start=offset + 1)),
                )
                candidates.append(table)
                break
        if len(candidates) != 1:
            raise IngestionError(
                f"{file_name}: expected exactly one classifiable worksheet, found {len(candidates)}"
            )
        return candidates[0]
    finally:
        workbook.close()


def _decode_delimited(payload: bytes) -> str:
    if payload.startswith((b"\xff\xfe", b"\xfe\xff")):
        return payload.decode("utf-16")
    if payload.startswith(b"\xef\xbb\xbf"):
        return payload.decode("utf-8-sig")
    try:
        return payload.decode("utf-8")
    except UnicodeDecodeError as exc:
        raise IngestionError("CSV encoding must be UTF-8 or BOM-marked UTF-16") from exc


def _read_delimited(file_name: str, payload: bytes) -> SourceTable:
    text = _decode_delimited(payload)
    delimiter = "\t" if "\t" in text[:4096] else ","
    raw_rows = [(line_number, row) for line_number, row in enumerate(csv.reader(io.StringIO(text), delimiter=delimiter), start=1)]
    header_position: int | None = None
    for position, (_, row) in enumerate(raw_rows[:20]):
        if sum(not _is_blank(value) for value in row) < 2:
            continue
        try:
            classify_headers(row)
        except IngestionError:
            continue
        header_position = position
        break
    if header_position is None:
        raise IngestionError(f"{file_name}: no recognized header row in first 20 lines")
    header_number, raw_headers = raw_rows[header_position]
    return _build_table(
        file_name=file_name,
        sheet_name="CSV",
        header_row_number=header_number,
        raw_headers=raw_headers,
        data_rows=raw_rows[header_position + 1 :],
    )


def read_source_payload(file_name: str, payload: bytes) -> SourceTable:
    """Classify one supported export from bytes without trusting its filename."""

    suffix = Path(file_name).suffix.casefold()
    if suffix not in SUPPORTED_SOURCE_SUFFIXES:
        supported = ", ".join(sorted(SUPPORTED_SOURCE_SUFFIXES))
        raise IngestionError(
            f"unsupported source file extension {suffix or '<none>'!r}; "
            f"expected one of {supported}"
        )
    return (
        _read_xlsx(file_name, payload)
        if suffix == ".xlsx"
        else _read_delimited(file_name, payload)
    )


def read_source_file(source_path: str | Path) -> SourceTable:
    """Read and classify one local export file by its header fingerprint."""

    path = Path(source_path)
    if not path.is_file():
        raise IngestionError(f"source must be a readable file: {path}")
    return read_source_payload(path.name, path.read_bytes())


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def ingest_bundle(bundle_path: str | Path) -> IngestedBundle:
    """Ingest the eight required exports from a ZIP without trusting file names."""

    path = Path(bundle_path)
    if not path.is_file() or not zipfile.is_zipfile(path):
        raise IngestionError(f"bundle must be a readable ZIP file: {path}")
    tables: Dict[str, SourceTable] = {}
    with zipfile.ZipFile(path) as archive:
        for member in archive.infolist():
            name = member.filename
            basename = PurePosixPath(name).name
            if member.is_dir() or name.startswith("__MACOSX/") or basename.startswith("._"):
                continue
            suffix = Path(basename).suffix.casefold()
            if suffix not in SUPPORTED_SOURCE_SUFFIXES:
                continue
            payload = archive.read(member)
            table = read_source_payload(basename, payload)
            if table.kind in tables:
                raise IngestionError(
                    f"duplicate source kind {table.kind}: {tables[table.kind].file_name}, {basename}"
                )
            tables[table.kind] = table
    missing = sorted(REQUIRED_SOURCE_KINDS - set(tables))
    unexpected = sorted(set(tables) - REQUIRED_SOURCE_KINDS)
    if missing or unexpected:
        raise IngestionError(f"bundle classification failed; missing={missing}, unexpected={unexpected}")
    return IngestedBundle(
        bundle_path=str(path.resolve()),
        bundle_sha256=_sha256(path),
        tables=tables,
    )


def row_value(row: SourceRow, header: str, *, required: bool = True) -> Any:
    """Read a value by normalized header, rejecting missing required columns."""

    wanted = normalize_header(header)
    for raw_header, value in row.values.items():
        if normalize_header(raw_header) == wanted:
            return value
    if required:
        raise IngestionError(f"row {row.row_number}: required header not found: {header}")
    return None


def header_position(table: SourceTable, header: str) -> Tuple[int, str]:
    wanted = normalize_header(header)
    for index, raw_header in enumerate(table.headers, start=1):
        if normalize_header(raw_header) == wanted:
            return index, raw_header
    raise IngestionError(f"{table.file_name}: required header not found: {header}")
